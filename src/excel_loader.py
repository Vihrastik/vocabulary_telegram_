from openpyxl import (
    load_workbook,
    Workbook,
)

from src.const import Language


class ExcelLoader:
    """
    Класс для работы с Excel из локальной директории.
    """

    wb = Workbook()
    files = {
        Language.eng: 'English_vocabulary.xlsx',
        Language.ger: 'Deutsch_vocabulary.xlsx',
        Language.fr: 'Français_vocabulary.xlsx'
    }

    def get_topics(self, lang):
        file = ExcelLoader.files.get(lang, None)
        if file is None:
            pass
        wb = load_workbook(filename=file)
        return wb.sheetnames

    def get_words(self, lang: str, topic: str) -> list:
        file = ExcelLoader.files.get(lang, None)
        wb = load_workbook(filename=file)
        sheet = wb.get_sheet_by_name(name=topic.capitalize())
        words = []
        for row in sheet.rows:
            if row[0].value is None or row[1].value is None:
                continue
            words.append([row[0].value, row[1].value])
        return words
